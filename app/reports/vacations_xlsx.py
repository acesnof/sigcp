"""Relatório Excel da gestão integrada de férias."""

from collections import Counter, defaultdict
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.vacation_service import ACTIONABLE_STATUSES, APPROVED_STATUSES


TEAL = "0B5962"
TEAL_DARK = "073F46"
TEAL_LIGHT = "DCEFF1"
RED = "B51618"
AMBER = "E5A93D"
GREEN = "2D7D61"
INK = "163438"
MUTED = "60777A"
WHITE = "FFFFFF"
LINE = "CAD8DA"


def _title(ws, title, subtitle, columns):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
    cell = ws.cell(1, 1, title)
    cell.font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=TEAL_DARK)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 34
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=columns)
    sub = ws.cell(2, 1, subtitle)
    sub.font = Font(name="Aptos", size=10, color=MUTED)
    sub.fill = PatternFill("solid", fgColor="F4F8F8")
    sub.alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 23


def _header(ws, row, labels):
    for column, label in enumerate(labels, 1):
        cell = ws.cell(row, column, label)
        cell.font = Font(name="Aptos", size=9, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="thin", color=TEAL_DARK))
    ws.row_dimensions[row].height = 28


def _body_style(ws, start_row, end_row, columns):
    thin = Side(style="hair", color=LINE)
    for row in range(start_row, end_row + 1):
        for col in range(1, columns + 1):
            cell = ws.cell(row, col)
            cell.font = Font(name="Aptos", size=9, color=INK)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
            if row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F7FAFA")


def _autosize(ws, minimum=9, maximum=42):
    for col in range(1, ws.max_column + 1):
        width = minimum
        for row in range(1, min(ws.max_row, 500) + 1):
            value = ws.cell(row, col).value
            if value is not None:
                width = max(width, min(maximum, len(str(value)) + 2))
        ws.column_dimensions[get_column_letter(col)].width = width


def _status_fill(status):
    if status in APPROVED_STATUSES:
        return GREEN
    if status in ACTIONABLE_STATUSES:
        return AMBER
    if status in ("Rejeitado", "Anulado"):
        return RED
    return MUTED


def _summary_sheet(workbook, payload):
    ws = workbook.active
    ws.title = "Resumo"
    _title(ws, "SIGCP · Gestão de Férias", f"Relatório consolidado · {payload['ano']}", 4)
    metrics = [
        ("Pessoas", payload["resumo"]["pessoas"]),
        ("Períodos planeados", payload["resumo"]["periodos"]),
        ("Períodos aprovados", payload["resumo"]["aprovados"]),
        ("Decisões pendentes", payload["resumo"]["pendentes"]),
        ("Dias de férias planeados", payload["resumo"]["dias_planeados"]),
        ("Dias ainda disponíveis", payload["resumo"]["dias_disponiveis"]),
    ]
    _header(ws, 4, ["Indicador", "Valor", "Indicador", "Valor"])
    for index in range(0, len(metrics), 2):
        row = 5 + index // 2
        left = metrics[index]
        right = metrics[index + 1] if index + 1 < len(metrics) else ("", "")
        ws.append([left[0], left[1], right[0], right[1]])
        for col in (1, 3):
            ws.cell(row, col).font = Font(name="Aptos", size=10, bold=True, color=TEAL_DARK)
        for col in (2, 4):
            ws.cell(row, col).font = Font(name="Aptos Display", size=15, bold=True, color=RED)

    status_counts = Counter(item["estado"] for item in payload["pedidos"])
    start = 10
    _header(ws, start, ["Estado", "Pedidos", "Área funcional", "Pedidos"])
    area_counts = Counter(item.get("area_funcional") or "Não definido" for item in payload["pedidos"])
    statuses = list(status_counts.items())
    areas = list(sorted(area_counts.items(), key=lambda item: item[0].casefold()))
    length = max(len(statuses), len(areas), 1)
    for index in range(length):
        row = start + 1 + index
        status_item = statuses[index] if index < len(statuses) else ("", "")
        area_item = areas[index] if index < len(areas) else ("", "")
        ws.cell(row, 1, status_item[0])
        ws.cell(row, 2, status_item[1])
        ws.cell(row, 3, area_item[0])
        ws.cell(row, 4, area_item[1])
        if status_item[0]:
            ws.cell(row, 1).font = Font(name="Aptos", size=9, bold=True, color=_status_fill(status_item[0]))
    _body_style(ws, 5, ws.max_row, 4)
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 14


def _requests_sheet(workbook, payload, *, pending_only=False):
    title = "Pendentes" if pending_only else "Pedidos"
    ws = workbook.create_sheet(title)
    rows = [
        item for item in payload["pedidos"]
        if not pending_only or item["estado"] in ACTIONABLE_STATUSES
    ]
    labels = [
        "ID", "Estado", "NIM", "Pessoa", "Área funcional", "Posição N.º", "Partida",
        "Chegada", "Companhia / voo", "F", "TD", "FS", "Observações",
        "Submetido em", "Decidido em", "Nota da decisão",
    ]
    _title(ws, f"{title} de férias", f"SIGCP · {payload['ano']}", len(labels))
    _header(ws, 4, labels)
    for item in rows:
        summary = item.get("resumo") or {}
        ws.append([
            item["id"], item["estado"], item.get("nim") or "",
            item.get("identificacao") or "", item.get("area_funcional") or "Não definido",
            item.get("posicao_numero") or "",
            item.get("data_hora_inicio") or "", item.get("data_hora_fim") or "",
            item.get("companhia_aerea") or "", summary.get("dias_ferias", 0),
            summary.get("dias_viagem", 0), summary.get("dias_fim_semana_feriado", 0),
            item.get("observacao") or "", item.get("submetido_em") or "",
            item.get("decidido_em") or "", item.get("nota_decisao") or "",
        ])
    if rows:
        _body_style(ws, 5, ws.max_row, len(labels))
        for row_index, item in enumerate(rows, 5):
            ws.cell(row_index, 2).font = Font(
                name="Aptos", size=9, bold=True, color=_status_fill(item["estado"])
            )
    ws.auto_filter.ref = f"A4:{get_column_letter(len(labels))}{max(4, ws.max_row)}"
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    _autosize(ws)


def _people_sheet(workbook, payload):
    ws = workbook.create_sheet("Pessoal")
    labels = [
        "NIM", "Pessoa", "Área funcional", "Posição N.º", "Antiguidade",
        "Telemóvel Serviço", "Início missão", "Fim missão", "Total dias Férias (manual)",
        "Direito", "Planeados", "Aprovados", "Disponíveis", "TD", "FS",
        "Períodos", "Pendentes", "Missão prorrogada",
    ]
    _title(ws, "Situação individual de férias", f"SIGCP · {payload['ano']}", len(labels))
    _header(ws, 4, labels)
    for person in payload["pessoas"]:
        summary = person["resumo"]
        ws.append([
            person["nim"], person["identificacao"], person["area_funcional"],
            person.get("posicao_numero") or "", person.get("antiguidade") or "",
            person.get("telemovel_servico") or "", person["data_chegada"],
            person["data_partida"], person.get("ferias_direito_override"), summary["direito"],
            summary["planeados"], summary["aprovados"], summary["disponiveis"],
            summary["dias_viagem"], summary["fins_semana_feriados"],
            summary["periodos"], summary["pendentes"],
            "Sim" if person["missao_prorrogada"] else "Não",
        ])
    if payload["pessoas"]:
        _body_style(ws, 5, ws.max_row, len(labels))
    ws.auto_filter.ref = f"A4:{get_column_letter(len(labels))}{max(4, ws.max_row)}"
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    _autosize(ws)


def _history_sheet(workbook, payload):
    ws = workbook.create_sheet("Histórico")
    labels = ["Pedido", "Pessoa", "Data/hora", "Ação", "Estado anterior", "Estado novo", "Autor", "Nota"]
    _title(ws, "Histórico e auditoria", f"SIGCP · {payload['ano']}", len(labels))
    _header(ws, 4, labels)
    events = []
    for item in payload["pedidos"]:
        for event in item.get("historico") or []:
            events.append((item, event))
    events.sort(key=lambda pair: str(pair[1].get("criado_em") or ""), reverse=True)
    for item, event in events:
        ws.append([
            item["id"], item.get("identificacao") or "", event.get("criado_em") or "",
            event.get("acao") or "", event.get("estado_anterior") or "",
            event.get("estado_novo") or "", event.get("ator") or "Sistema",
            event.get("nota") or "",
        ])
    if events:
        _body_style(ws, 5, ws.max_row, len(labels))
    ws.auto_filter.ref = f"A4:{get_column_letter(len(labels))}{max(4, ws.max_row)}"
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    _autosize(ws)


def _holidays_sheet(workbook, payload):
    ws = workbook.create_sheet("Feriados")
    _title(ws, "Feriados considerados", f"SIGCP · {payload['ano']}", 3)
    _header(ws, 4, ["Data", "Descrição", "Ativo"])
    for holiday in payload.get("feriados") or []:
        ws.append([holiday["data"], holiday["descricao"], "Sim" if holiday["ativo"] else "Não"])
    if payload.get("feriados"):
        _body_style(ws, 5, ws.max_row, 3)
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    _autosize(ws)


def generate_vacations_xlsx(payload):
    workbook = Workbook()
    workbook.properties.title = f"SIGCP - Férias {payload['ano']}"
    workbook.properties.subject = "Gestão de férias do Contingente Português"
    workbook.properties.creator = "SIGCP"
    _summary_sheet(workbook, payload)
    _requests_sheet(workbook, payload)
    _people_sheet(workbook, payload)
    _requests_sheet(workbook, payload, pending_only=True)
    _history_sheet(workbook, payload)
    _holidays_sheet(workbook, payload)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
