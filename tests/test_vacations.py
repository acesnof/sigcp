import json
import sqlite3
import tempfile
import unittest
from datetime import date, datetime, timedelta
from pathlib import Path
from unittest.mock import patch

from app import db, vacation_service
from app.config import MASTER_NIM, MASTER_PASSWORD
from app.security import hash_password
from app.web_app import create_web_app


class VacationWorkflowTest(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.db_path = Path(self.temp_dir.name) / "vacations.sqlite3"
        self.db_patch = patch.object(db, "DB_PATH", str(self.db_path))
        self.db_patch.start()
        db.init_db()
        self.person_id = self.create_user("militar", area="HQ")
        self.snr_id = self.create_user("snr_ferias", snr=1, area="HQ")
        self.app = create_web_app()
        self.app.config.update(TESTING=True)
        self.client = self.app.test_client()

    def tearDown(self):
        self.db_patch.stop()
        self.temp_dir.cleanup()

    def create_user(
        self,
        nim,
        *,
        snr=0,
        area="Não definido",
        posto="OR-5",
        antiguidade="2020-01-01",
        posicao_numero="",
        partida="2026-12-31 20:00",
        acesso="Leitura",
    ):
        salt, pwd_hash = hash_password("Teste123!")
        return db.db_execute_return_id(
            """
            INSERT INTO utilizadores (
                nim, posto, nome, sobrenome, data_chegada, data_partida,
                antiguidade, snr, telemovel_servico, responsavel_welfare,
                tipo_acesso, password_salt, password_hash, master,
                area_funcional, posicao_numero
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, '', 0, ?, ?, ?, 0, ?, ?)
            """,
            (
                nim,
                posto,
                "Nome",
                nim,
                "2026-01-01 08:00",
                partida,
                antiguidade,
                snr,
                acesso,
                salt,
                pwd_hash,
                area,
                posicao_numero,
            ),
        )

    def test_authorized_profiles_update_only_hours_of_approved_vacations(self):
        updater_id = self.create_user(
            "gestor_horas", acesso="Gestão Welfare Individual"
        )
        _boot, person_headers = self.login("militar")
        vacation_id = self.create_request(person_headers)
        self.logout(person_headers)
        _boot, snr_headers = self.login("snr_ferias")
        approved = self.client.post(
            f"/api/vacations/{vacation_id}/decision",
            json={"action": "approve", "note": ""},
            headers=snr_headers,
        )
        self.assertEqual(200, approved.status_code, approved.get_json())
        self.logout(snr_headers)

        boot, updater_headers = self.login("gestor_horas")
        self.assertTrue(boot["permissions"]["ferias"])
        self.assertFalse(boot["permissions"]["ferias_gerir"])
        self.assertTrue(boot["permissions"]["ferias_atualizar_horas"])
        management = self.client.get("/api/vacations/manage?ano=2026")
        self.assertEqual(200, management.status_code)
        updated = self.client.put(
            f"/api/vacations/{vacation_id}/hours",
            json={"hora_partida": "20:15", "hora_chegada": "08:45"},
            headers=updater_headers,
        )
        self.assertEqual(200, updated.status_code, updated.get_json())
        row = db.db_one("SELECT * FROM ferias WHERE id=?", (vacation_id,))
        self.assertEqual("2026-06-10 20:15", row["data_hora_inicio"])
        self.assertEqual("2026-06-20 08:45", row["data_hora_fim"])
        self.assertEqual("Aprovado", row["estado"])
        history = db.db_one(
            "SELECT utilizador_id, acao FROM ferias_historico WHERE feria_id=? ORDER BY id DESC LIMIT 1",
            (vacation_id,),
        )
        self.assertEqual(updater_id, history["utilizador_id"])
        self.assertEqual("Horas atualizadas", history["acao"])

    def login(self, nim, password="Teste123!"):
        response = self.client.post(
            "/api/login", json={"nim": nim, "password": password}
        )
        self.assertEqual(200, response.status_code)
        boot = self.client.get("/api/bootstrap").get_json()
        return boot, {"X-CSRF-Token": boot["csrf_token"]}

    def logout(self, headers):
        self.client.post("/api/logout", headers=headers)

    def create_request(self, headers, **overrides):
        payload = {
            "data_hora_inicio": "2026-06-10T18:00",
            "data_hora_fim": "2026-06-20T10:00",
            "companhia_aerea": "TAP TP123",
            "observacao": "Licença anual",
            "accept_warnings": True,
        }
        payload.update(overrides)
        response = self.client.post("/api/vacations", json=payload, headers=headers)
        self.assertEqual(200, response.status_code, response.get_json())
        return response.get_json()["id"]

    def test_private_area_snr_management_approval_and_welfare_reflection(self):
        boot, person_headers = self.login("militar")
        self.assertTrue(boot["permissions"]["ferias_privadas"])
        self.assertFalse(boot["permissions"]["ferias"])

        warning = self.client.post(
            "/api/vacations",
            json={
                "data_hora_inicio": "2026-06-10T18:00",
                "data_hora_fim": "2026-06-20T10:00",
            },
            headers=person_headers,
        )
        self.assertEqual(409, warning.status_code)
        self.assertTrue(warning.get_json()["warnings"])

        vacation_id = self.create_request(person_headers)
        own = self.client.get("/api/vacations/me?ano=2026&todos=1").get_json()["data"]
        self.assertEqual("Pendente", own["pedidos"][0]["estado"])
        self.assertEqual([], own["pedidos"][0]["historico"])
        self.assertGreaterEqual(own["resumo"]["planeados"], 1)
        detail = self.client.get(f"/api/vacations/{vacation_id}").get_json()["data"]
        self.assertTrue(detail["historico"])
        self.logout(person_headers)

        boot, snr_headers = self.login("snr_ferias")
        self.assertTrue(boot["permissions"]["ferias"])
        self.assertTrue(boot["permissions"]["ferias_decidir"])
        self.assertTrue(boot["permissions"]["pessoal"])
        self.assertFalse(boot["permissions"]["pessoal_editar"])
        management = self.client.get("/api/vacations/manage?ano=2026")
        self.assertEqual(200, management.status_code)
        self.assertEqual(1, management.get_json()["data"]["resumo"]["pendentes"])

        approved = self.client.post(
            f"/api/vacations/{vacation_id}/decision",
            json={"action": "approve", "note": "Autorizado"},
            headers=snr_headers,
        )
        self.assertEqual(200, approved.status_code, approved.get_json())
        self.assertEqual(
            "Aprovado",
            db.db_one("SELECT estado FROM ferias WHERE id=?", (vacation_id,))["estado"],
        )
        self.assertIn(self.person_id, db.get_ferias_mes(2026, 6))

        self.logout(snr_headers)
        master = self.client.post(
            "/api/login", json={"nim": MASTER_NIM, "password": MASTER_PASSWORD}
        )
        self.assertEqual(200, master.status_code)
        individual = self.client.get(
            "/api/individual?ano=2026&mes=6&modo=welfare"
        ).get_json()["data"]
        row = next(item for item in individual["linhas"] if item["id"] == self.person_id)
        self.assertEqual("ferias", row["celulas"]["2026-06-12"]["almoco"]["estado"])
        self.assertEqual("ferias", row["celulas"]["2026-06-12"]["jantar"]["estado"])

    def test_change_and_cancellation_keep_approved_period_until_decision(self):
        _boot, person_headers = self.login("militar")
        vacation_id = self.create_request(person_headers)
        self.logout(person_headers)
        _boot, snr_headers = self.login("snr_ferias")
        self.client.post(
            f"/api/vacations/{vacation_id}/decision",
            json={"action": "approve", "note": ""},
            headers=snr_headers,
        )
        self.logout(snr_headers)

        _boot, person_headers = self.login("militar")
        changed = self.client.post(
            f"/api/vacations/{vacation_id}/change-request",
            json={
                "data_hora_inicio": "2026-07-10T18:00",
                "data_hora_fim": "2026-07-18T10:00",
                "reason": "Alteração do voo",
                "accept_warnings": True,
            },
            headers=person_headers,
        )
        self.assertEqual(200, changed.status_code, changed.get_json())
        pending_change = db.db_one("SELECT * FROM ferias WHERE id=?", (vacation_id,))
        self.assertEqual("Alteração pendente", pending_change["estado"])
        self.assertEqual("2026-06-10 18:00", pending_change["data_hora_inicio"])
        self.assertEqual("2026-07-10 18:00", pending_change["proposta_data_hora_inicio"])
        self.assertIn(self.person_id, db.get_ferias_mes(2026, 6))
        self.logout(person_headers)

        _boot, snr_headers = self.login("snr_ferias")
        decision = self.client.post(
            f"/api/vacations/{vacation_id}/change-decision",
            json={"action": "approve", "note": ""},
            headers=snr_headers,
        )
        self.assertEqual(200, decision.status_code, decision.get_json())
        accepted_change = db.db_one("SELECT * FROM ferias WHERE id=?", (vacation_id,))
        self.assertEqual("Aprovado", accepted_change["estado"])
        self.assertEqual("2026-07-10 18:00", accepted_change["data_hora_inicio"])
        self.assertNotIn(self.person_id, db.get_ferias_mes(2026, 6))
        self.assertIn(self.person_id, db.get_ferias_mes(2026, 7))
        self.logout(snr_headers)

        _boot, person_headers = self.login("militar")
        cancellation = self.client.post(
            f"/api/vacations/{vacation_id}/cancellation-request",
            json={"reason": "Necessidade de serviço"},
            headers=person_headers,
        )
        self.assertEqual(200, cancellation.status_code)
        self.assertIn(self.person_id, db.get_ferias_mes(2026, 7))
        self.logout(person_headers)

        _boot, snr_headers = self.login("snr_ferias")
        cancelled = self.client.post(
            f"/api/vacations/{vacation_id}/cancellation-decision",
            json={"action": "approve", "note": ""},
            headers=snr_headers,
        )
        self.assertEqual(200, cancelled.status_code, cancelled.get_json())
        self.assertEqual(
            "Anulado",
            db.db_one("SELECT estado FROM ferias WHERE id=?", (vacation_id,))["estado"],
        )
        self.assertNotIn(self.person_id, db.get_ferias_mes(2026, 7))

    def test_separation_of_duties_calendar_notifications_and_excel(self):
        _boot, snr_headers = self.login("snr_ferias")
        own_id = self.create_request(snr_headers)
        self.assertEqual(
            "Aprovado",
            db.db_one("SELECT estado FROM ferias WHERE id=?", (own_id,))["estado"],
        )
        self.assertEqual(
            self.snr_id,
            db.db_one("SELECT decidido_por FROM ferias WHERE id=?", (own_id,))["decidido_por"],
        )
        change = self.client.post(
            f"/api/vacations/{own_id}/change-request",
            json={
                "data_hora_inicio": "2026-06-11T18:00",
                "data_hora_fim": "2026-06-21T10:00",
                "reason": "Alterar datas",
                "accept_warnings": True,
            },
            headers=snr_headers,
        )
        self.assertEqual(200, change.status_code, change.get_json())
        changed_row = db.db_one("SELECT * FROM ferias WHERE id=?", (own_id,))
        self.assertEqual("Aprovado", changed_row["estado"])
        self.assertEqual("2026-06-11 18:00", changed_row["data_hora_inicio"])
        self.assertEqual(self.snr_id, changed_row["decidido_por"])

        db.db_execute(
            "UPDATE utilizadores SET posto='OR-9', sobrenome='FONSECA', data_nascimento='1980-06-15' WHERE id=?",
            (self.snr_id,),
        )
        monthly_calendar = self.client.get("/api/calendar?ano=2026&mes=6")
        self.assertEqual(200, monthly_calendar.status_code)
        self.assertEqual(
            "OR-9 FONSECA",
            monthly_calendar.get_json()["aniversarios"]["2026-06-15"][0]["identificacao"],
        )

        calendar = self.client.get(
            "/api/vacations/calendar?ano=2026&mes=6&scope=all"
        )
        self.assertEqual(200, calendar.status_code)
        self.assertIn("grelha", calendar.get_json()["data"])

        report = self.client.get("/api/vacations/report.xlsx?ano=2026")
        self.assertEqual(200, report.status_code)
        self.assertEqual(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            report.mimetype,
        )
        self.assertIn(
            "SIGCP_Ferias_2026.xlsx",
            report.headers.get("Content-Disposition", ""),
        )
        self.assertGreater(len(report.data), 2000)

        read = self.client.post(
            "/api/vacations/notifications/read", json={}, headers=snr_headers
        )
        self.assertEqual(200, read.status_code)
        unread = db.db_one(
            "SELECT COUNT(*) AS total FROM ferias_notificacoes WHERE utilizador_id=? AND lida=0",
            (self.snr_id,),
        )["total"]
        self.assertEqual(0, unread)

    def test_management_and_excel_only_include_departed_people_when_requested(self):
        departed_id = self.create_user(
            "departed", partida="2026-01-31 20:00"
        )
        _boot, snr_headers = self.login("snr_ferias")

        current = self.client.get("/api/vacations/manage?ano=2026")
        self.assertEqual(200, current.status_code)
        current_data = current.get_json()["data"]
        self.assertEqual(2, current_data["resumo"]["pessoas"])
        self.assertIn(departed_id, [item["id"] for item in current_data["pessoas"]])

        all_people = self.client.get("/api/vacations/manage?ano=2026&todos=1")
        self.assertEqual(200, all_people.status_code)
        all_data = all_people.get_json()["data"]
        self.assertEqual(3, all_data["resumo"]["pessoas"])
        self.assertIn(departed_id, [item["id"] for item in all_data["pessoas"]])

        from openpyxl import load_workbook
        from io import BytesIO

        current_report = self.client.get("/api/vacations/report.xlsx?ano=2026&todos=0")
        all_report = self.client.get("/api/vacations/report.xlsx?ano=2026&todos=1")
        current_book = load_workbook(BytesIO(current_report.data), read_only=True)
        all_book = load_workbook(BytesIO(all_report.data), read_only=True)
        self.assertEqual(2, current_book["Resumo"]["B5"].value)
        self.assertEqual(3, all_book["Resumo"]["B5"].value)
        current_names = {
            row[1] for row in current_book["Pessoal"].iter_rows(min_row=5, values_only=True)
        }
        all_names = {
            row[1] for row in all_book["Pessoal"].iter_rows(min_row=5, values_only=True)
        }
        self.assertNotIn("OR-5 DEPARTED", current_names)
        self.assertIn("OR-5 DEPARTED", all_names)

    def test_current_view_history_and_chronological_order(self):
        now = datetime.now().replace(second=0, microsecond=0)
        periods = [
            (now - timedelta(days=20), now - timedelta(days=10), "passado"),
            (now + timedelta(days=12), now + timedelta(days=18), "futuro dois"),
            (now + timedelta(days=3), now + timedelta(days=8), "futuro um longo"),
            (now + timedelta(days=3), now + timedelta(days=6), "futuro um curto"),
        ]
        ids = []
        for start, end, note in periods:
            ids.append(
                db.db_execute_return_id(
                    """
                    INSERT INTO ferias (
                        utilizador_id, data_hora_inicio, data_hora_fim,
                        observacao, estado, submetido_por, submetido_em
                    ) VALUES (?, ?, ?, ?, 'Aprovado', ?, CURRENT_TIMESTAMP)
                    """,
                    (
                        self.person_id,
                        start.strftime("%Y-%m-%d %H:%M"),
                        end.strftime("%Y-%m-%d %H:%M"),
                        note,
                        self.person_id,
                    ),
                )
            )

        _boot, headers = self.login("militar")
        current = self.client.get("/api/vacations/me?ano=2026").get_json()["data"]
        self.assertFalse(current["mostrar_tudo"])
        self.assertEqual(
            [ids[3], ids[2], ids[1]],
            [item["id"] for item in current["pedidos"]],
        )

        history = self.client.get("/api/vacations/me?ano=2026&todos=1").get_json()["data"]
        self.assertTrue(history["mostrar_tudo"])
        self.assertEqual(
            [ids[0], ids[3], ids[2], ids[1]],
            [item["id"] for item in history["pedidos"]],
        )
        self.logout(headers)

        _boot, _headers = self.login("snr_ferias")
        managed = self.client.get("/api/vacations/manage?ano=2026").get_json()["data"]
        self.assertEqual(
            [ids[3], ids[2], ids[1]],
            [item["id"] for item in managed["pedidos"]],
        )
        managed_history = self.client.get(
            "/api/vacations/manage?ano=2026&todos=1"
        ).get_json()["data"]
        self.assertEqual(
            [ids[0], ids[3], ids[2], ids[1]],
            [item["id"] for item in managed_history["pedidos"]],
        )

    def test_snr_and_admin_restore_annulled_vacations(self):
        vacation_id = db.db_execute_return_id(
            """
            INSERT INTO ferias (
                utilizador_id, data_hora_inicio, data_hora_fim, observacao,
                estado, submetido_por, submetido_em, motivo_anulacao
            ) VALUES (?, '2026-09-10 18:00', '2026-09-20 10:00',
                      'Restaurar', 'Anulado', ?, CURRENT_TIMESTAMP, 'Teste')
            """,
            (self.person_id, self.person_id),
        )
        db.db_execute(
            """
            INSERT INTO ferias_historico (
                feria_id, utilizador_id, acao, estado_anterior, estado_novo
            ) VALUES (?, ?, 'Anulado', 'Aprovado', 'Anulado')
            """,
            (vacation_id, self.snr_id),
        )

        _boot, person_headers = self.login("militar")
        denied = self.client.post(
            f"/api/vacations/{vacation_id}/restore", json={}, headers=person_headers
        )
        self.assertEqual(403, denied.status_code)
        self.logout(person_headers)

        _boot, snr_headers = self.login("snr_ferias")
        restored = self.client.post(
            f"/api/vacations/{vacation_id}/restore", json={}, headers=snr_headers
        )
        self.assertEqual(200, restored.status_code, restored.get_json())
        self.assertEqual(
            "Aprovado",
            db.db_one("SELECT estado FROM ferias WHERE id=?", (vacation_id,))["estado"],
        )
        self.logout(snr_headers)

        db.db_execute(
            "UPDATE ferias SET estado='Anulado', motivo_anulacao='Segundo teste' WHERE id=?",
            (vacation_id,),
        )
        db.db_execute(
            """
            INSERT INTO ferias_historico (
                feria_id, utilizador_id, acao, estado_anterior, estado_novo
            ) VALUES (?, ?, 'Anulado', 'Aprovado', 'Anulado')
            """,
            (vacation_id, self.snr_id),
        )
        _boot, admin_headers = self.login(MASTER_NIM, MASTER_PASSWORD)
        restored = self.client.post(
            f"/api/vacations/{vacation_id}/restore", json={}, headers=admin_headers
        )
        self.assertEqual(200, restored.status_code, restored.get_json())
        self.assertEqual(
            "Aprovado",
            db.db_one("SELECT estado FROM ferias WHERE id=?", (vacation_id,))["estado"],
        )

    def test_notifications_can_be_read_deleted_and_are_private(self):
        own_notification = db.db_execute_return_id(
            """
            INSERT INTO ferias_notificacoes (
                utilizador_id, tipo, titulo, mensagem, lida
            ) VALUES (?, 'teste', 'Aviso', 'Mensagem', 0)
            """,
            (self.person_id,),
        )
        other_notification = db.db_execute_return_id(
            """
            INSERT INTO ferias_notificacoes (
                utilizador_id, tipo, titulo, mensagem, lida
            ) VALUES (?, 'teste', 'Outro', 'Privada', 0)
            """,
            (self.snr_id,),
        )
        _boot, headers = self.login("militar")
        notification_list = self.client.get("/api/vacations/notifications")
        self.assertEqual(200, notification_list.status_code)
        notification_data = notification_list.get_json()["data"]
        self.assertEqual(1, notification_data["nao_lidas"])
        self.assertEqual(own_notification, notification_data["notificacoes"][0]["id"])
        read = self.client.post(
            "/api/vacations/notifications/read",
            json={"id": own_notification},
            headers=headers,
        )
        self.assertEqual(200, read.status_code)
        self.assertEqual(
            1,
            db.db_one(
                "SELECT lida FROM ferias_notificacoes WHERE id=?", (own_notification,)
            )["lida"],
        )
        denied = self.client.delete(
            f"/api/vacations/notifications/{other_notification}", headers=headers
        )
        self.assertEqual(404, denied.status_code)
        deleted = self.client.delete(
            f"/api/vacations/notifications/{own_notification}", headers=headers
        )
        self.assertEqual(200, deleted.status_code)
        self.assertIsNone(
            db.db_one("SELECT id FROM ferias_notificacoes WHERE id=?", (own_notification,))
        )

    def test_management_page_has_bounded_database_round_trips(self):
        for index in range(12):
            self.create_user(f"rede_{index:02d}", area="OPS")
        _boot, _headers = self.login("snr_ferias")

        with patch.object(db, "_connect", wraps=db._connect) as connect:
            response = self.client.get("/api/vacations/manage?ano=2026")

        self.assertEqual(200, response.status_code)
        self.assertLessEqual(
            connect.call_count,
            8,
            "A Gestão de Férias voltou a abrir a base de dados por pessoa.",
        )

    def test_new_managed_request_only_marks_people_still_in_mission(self):
        departed_id = self.create_user("missao_terminada", area="OPS")
        future_id = self.create_user("missao_ativa", area="OPS")
        no_departure_id = self.create_user("missao_sem_fim", area="OPS")
        db.db_execute(
            "UPDATE utilizadores SET data_partida=? WHERE id=?",
            ((datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d %H:%M"), departed_id),
        )
        db.db_execute(
            "UPDATE utilizadores SET data_partida=? WHERE id=?",
            ((datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d %H:%M"), future_id),
        )
        db.db_execute(
            "UPDATE utilizadores SET data_partida='' WHERE id=?",
            (no_departure_id,),
        )
        _boot, _headers = self.login("snr_ferias")

        response = self.client.get("/api/vacations/manage?ano=2026")

        self.assertEqual(200, response.status_code)
        people = {
            item["id"]: item
            for item in response.get_json()["data"]["pessoas"]
        }
        self.assertFalse(people[departed_id]["pode_novo_pedido"])
        self.assertTrue(people[future_id]["pode_novo_pedido"])
        self.assertTrue(people[no_departure_id]["pode_novo_pedido"])

    def test_mission_eligibility_uses_the_departure_time(self):
        reference = datetime(2026, 7, 31, 12, 0)
        self.assertFalse(
            vacation_service.member_still_in_mission(
                {"data_partida": "2026-07-31 11:59"}, reference
            )
        )
        self.assertTrue(
            vacation_service.member_still_in_mission(
                {"data_partida": "2026-07-31 12:01"}, reference
            )
        )
        self.assertTrue(
            vacation_service.member_still_in_mission(
                {"data_partida": ""}, reference
            )
        )

    def test_calendar_orders_by_rank_then_antiquity(self):
        same_date_officer = self.create_user(
            "cal_antigo_of", posto="OF-2", antiguidade="2018-01-01"
        )
        same_date_soldier = self.create_user(
            "cal_antigo_or", posto="OR-1", antiguidade="2018-01-01"
        )
        newer_high_rank = self.create_user(
            "cal_recente_of", posto="OF-6", antiguidade="2020-01-01"
        )
        no_date_officer = self.create_user(
            "cal_sem_data_of", posto="OF-1", antiguidade=""
        )
        no_date_sergeant = self.create_user(
            "cal_sem_data_or", posto="OR-9", antiguidade=""
        )

        payload = vacation_service.calendar_payload(2026, 8)
        relevant = {
            same_date_officer,
            same_date_soldier,
            newer_high_rank,
            no_date_officer,
            no_date_sergeant,
        }
        order = [
            person["id"] for person in payload["pessoas"]
            if person["id"] in relevant
        ]

        self.assertEqual(
            [
                newer_high_rank,
                same_date_officer,
                no_date_officer,
                no_date_sergeant,
                same_date_soldier,
            ],
            order,
        )

    def test_print_payload_orders_by_rank_then_antiquity_and_lists_planned_periods(self):
        older_id = self.create_user(
            "print_antigo", antiguidade="2018-01-01", posicao_numero="A-01",
            partida="2099-12-31 20:00",
        )
        newer_id = self.create_user(
            "print_recente", antiguidade="2022-01-01", posicao_numero="A-02",
            partida="2099-12-31 20:00",
        )
        no_date_officer_id = self.create_user(
            "print_sem_data_of", posto="OF-2", antiguidade="", posicao_numero="B-01",
            partida="2099-12-31 20:00",
        )
        no_date_sergeant_id = self.create_user(
            "print_sem_data_or", posto="OR-8", antiguidade="", posicao_numero="B-02",
            partida="2099-12-31 20:00",
        )
        departed_id = self.create_user(
            "print_fora_missao", posto="OF-6", antiguidade="2010-01-01",
            partida="2020-01-01 08:00",
        )
        approved_id = db.db_execute_return_id(
            """
            INSERT INTO ferias (
                utilizador_id, data_hora_inicio, data_hora_fim, estado
            ) VALUES (?, '2026-08-10 18:00', '2026-08-20 10:00', 'Aprovado')
            """,
            (older_id,),
        )
        rejected_id = db.db_execute_return_id(
            """
            INSERT INTO ferias (
                utilizador_id, data_hora_inicio, data_hora_fim, estado
            ) VALUES (?, '2026-09-10 18:00', '2026-09-20 10:00', 'Rejeitado')
            """,
            (older_id,),
        )
        departed_period_id = db.db_execute_return_id(
            """
            INSERT INTO ferias (
                utilizador_id, data_hora_inicio, data_hora_fim, estado
            ) VALUES (?, '2026-10-10 18:00', '2026-10-20 10:00', 'Aprovado')
            """,
            (departed_id,),
        )
        for offset, person_id in enumerate(
            (newer_id, no_date_officer_id, no_date_sergeant_id), start=1
        ):
            db.db_execute_return_id(
                """
                INSERT INTO ferias (
                    utilizador_id, data_hora_inicio, data_hora_fim, estado
                ) VALUES (?, ?, ?, 'Aprovado')
                """,
                (
                    person_id,
                    f"2026-0{offset + 1}-10 18:00",
                    f"2026-0{offset + 1}-15 10:00",
                ),
            )
        pending_id = db.db_execute_return_id(
            """
            INSERT INTO ferias (
                utilizador_id, data_hora_inicio, data_hora_fim, estado
            ) VALUES (?, '2026-11-01 18:00', '2026-11-05 10:00', 'Pendente')
            """,
            (older_id,),
        )
        annulled_id = db.db_execute_return_id(
            """
            INSERT INTO ferias (
                utilizador_id, data_hora_inicio, data_hora_fim, estado
            ) VALUES (?, '2026-12-01 18:00', '2026-12-05 10:00', 'Anulado')
            """,
            (newer_id,),
        )

        payload = vacation_service.management_payload(year=2026, show_all=True)

        relevant_ids = {
            older_id, newer_id, no_date_officer_id, no_date_sergeant_id
        }
        ordered_relevant = [
            person_id
            for person_id in payload["ordem_impressao"]
            if person_id in relevant_ids
        ]
        self.assertEqual(
            [no_date_officer_id, no_date_sergeant_id, older_id, newer_id],
            ordered_relevant,
        )
        self.assertNotIn(departed_id, payload["ordem_impressao"])
        printed_ids = {item["id"] for item in payload["periodos_impressao"]}
        self.assertIn(approved_id, printed_ids)
        self.assertIn(rejected_id, printed_ids)
        self.assertIn(pending_id, printed_ids)
        self.assertIn(annulled_id, printed_ids)
        self.assertNotIn(departed_period_id, printed_ids)
        pending_payload = vacation_service.management_payload(
            year=2026, show_all=True, status_group="pending"
        )
        self.assertEqual(
            "Licenças pendentes de aprovação",
            pending_payload["titulo_impressao"],
        )
        self.assertEqual(
            {"Pendente", "Alteração pendente", "Cancelamento pendente"},
            set(vacation_service.STATUS_FILTER_GROUPS["pending"]),
        )
        self.assertEqual(
            {pending_id},
            {item["id"] for item in pending_payload["periodos_impressao"]},
        )
        approved_payload = vacation_service.management_payload(
            year=2026, show_all=True, status_group="approved"
        )
        self.assertTrue(approved_payload["periodos_impressao"])
        self.assertEqual(
            {"Aprovado"},
            {item["estado"] for item in approved_payload["periodos_impressao"]},
        )
        annulled_payload = vacation_service.management_payload(
            year=2026, show_all=True, status_group="annulled"
        )
        self.assertEqual(
            {annulled_id},
            {item["id"] for item in annulled_payload["periodos_impressao"]},
        )
        person = next(item for item in payload["pessoas"] if item["id"] == older_id)
        self.assertEqual("A-01", person["posicao_numero"])

    def test_admin_can_permanently_delete_vacations_in_every_status(self):
        _boot, admin_headers = self.login(MASTER_NIM, MASTER_PASSWORD)
        for index, status in enumerate(vacation_service.ALL_STATUSES, start=1):
            vacation_id = db.db_execute_return_id(
                """
                INSERT INTO ferias (
                    utilizador_id, data_hora_inicio, data_hora_fim, estado
                ) VALUES (?, ?, ?, ?)
                """,
                (
                    self.person_id,
                    f"2026-10-{index:02d} 18:00",
                    f"2026-10-{index + 1:02d} 10:00",
                    status,
                ),
            )
            db.db_execute(
                """
                INSERT INTO ferias_historico (
                    feria_id, utilizador_id, acao, estado_novo
                ) VALUES (?, ?, 'Teste', ?)
                """,
                (vacation_id, self.person_id, status),
            )
            db.db_execute(
                """
                INSERT INTO ferias_notificacoes (
                    utilizador_id, feria_id, tipo, canal, titulo
                ) VALUES (?, ?, 'teste', 'pessoal', 'Teste')
                """,
                (self.person_id, vacation_id),
            )
            deleted = self.client.delete(
                f"/api/vacations/{vacation_id}", headers=admin_headers
            )
            self.assertEqual(200, deleted.status_code, deleted.get_json())
            self.assertIsNone(db.db_one("SELECT id FROM ferias WHERE id=?", (vacation_id,)))
            self.assertIsNone(
                db.db_one("SELECT id FROM ferias_historico WHERE feria_id=?", (vacation_id,))
            )
            self.assertIsNone(
                db.db_one("SELECT id FROM ferias_notificacoes WHERE feria_id=?", (vacation_id,))
            )
        audit_rows = db.db_rows(
            "SELECT * FROM auditoria WHERE acao=?",
            ("Período de férias apagado definitivamente",),
        )
        self.assertEqual(len(vacation_service.ALL_STATUSES), len(audit_rows))

    def test_snr_management_notifications_are_separate_from_personal_ones(self):
        _boot, person_headers = self.login("militar")
        other_person_request = self.create_request(person_headers)
        self.logout(person_headers)

        snr_boot, snr_headers = self.login("snr_ferias")
        self.assertEqual(0, snr_boot["notifications"]["ferias_pessoais_nao_lidas"])
        self.assertEqual(1, snr_boot["notifications"]["ferias_gestao_nao_lidas"])
        personal = self.client.get(
            "/api/vacations/notifications?canal=pessoal"
        ).get_json()["data"]
        management = self.client.get(
            "/api/vacations/notifications?canal=gestao"
        ).get_json()["data"]
        self.assertEqual([], personal["notificacoes"])
        self.assertEqual(other_person_request, management["notificacoes"][0]["feria_id"])
        self.client.post(
            "/api/vacations/notifications/read",
            json={"canal": "gestao"},
            headers=snr_headers,
        )
        own_request = self.create_request(
            snr_headers,
            data_hora_inicio="2026-07-10T18:00",
            data_hora_fim="2026-07-15T10:00",
        )
        self.assertEqual(
            "Aprovado",
            db.db_one("SELECT estado FROM ferias WHERE id=?", (own_request,))["estado"],
        )
        self.logout(snr_headers)

        admin_boot, admin_headers = self.login(MASTER_NIM, MASTER_PASSWORD)
        self.assertEqual(0, admin_boot["notifications"]["ferias_gestao_nao_lidas"])
        denied_management = self.client.get(
            "/api/vacations/notifications?canal=gestao"
        )
        self.assertEqual(403, denied_management.status_code)
        already_decided = self.client.post(
            f"/api/vacations/{own_request}/decision",
            json={"action": "approve", "note": "Autorizado"},
            headers=admin_headers,
        )
        self.assertEqual(400, already_decided.status_code)
        self.logout(admin_headers)

        separated_boot, _headers = self.login("snr_ferias")
        self.assertEqual(
            0, separated_boot["notifications"]["ferias_pessoais_nao_lidas"]
        )
        self.assertEqual(
            0, separated_boot["notifications"]["ferias_gestao_nao_lidas"]
        )

    def test_active_snr_substitute_receives_new_management_notifications(self):
        substitute_id = self.create_user("snr_substituto_ferias", area="HQ")
        today = date.today().isoformat()
        db.db_execute(
            """
            UPDATE utilizadores
            SET snr_substituto=1, snr_substituto_inicio=?, snr_substituto_fim=?
            WHERE id=?
            """,
            (today, today, substitute_id),
        )
        self.assertEqual(
            substitute_id, db.get_snr_unico_para_assinatura()["id"]
        )

        _boot, person_headers = self.login("militar")
        vacation_id = self.create_request(person_headers)
        self.logout(person_headers)

        titular_boot, titular_headers = self.login("snr_ferias")
        self.assertEqual(
            0, titular_boot["notifications"]["ferias_gestao_nao_lidas"]
        )
        self.logout(titular_headers)

        substitute_boot, substitute_headers = self.login("snr_substituto_ferias")
        self.assertTrue(substitute_boot["permissions"]["snr"])
        self.assertTrue(substitute_boot["permissions"]["ferias_decidir"])
        self.assertEqual(
            1, substitute_boot["notifications"]["ferias_gestao_nao_lidas"]
        )
        notifications = self.client.get(
            "/api/vacations/notifications?canal=gestao"
        ).get_json()["data"]["notificacoes"]
        self.assertEqual(vacation_id, notifications[0]["feria_id"])
        self.logout(substitute_headers)


class VacationMigrationTest(unittest.TestCase):
    def test_legacy_period_is_preserved_and_marked_approved(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "legacy.sqlite3"
            salt, password_hash = hash_password("Teste123!")
            connection = sqlite3.connect(path)
            connection.executescript(
                """
                CREATE TABLE utilizadores (
                    id INTEGER PRIMARY KEY AUTOINCREMENT, nim TEXT NOT NULL UNIQUE,
                    posto TEXT, nome TEXT, sobrenome TEXT, data_chegada TEXT,
                    data_partida TEXT, antiguidade TEXT, snr INTEGER DEFAULT 0,
                    telemovel_servico TEXT, responsavel_welfare INTEGER DEFAULT 0,
                    tipo_acesso TEXT NOT NULL, password_salt TEXT NOT NULL,
                    password_hash TEXT NOT NULL, master INTEGER DEFAULT 0,
                    criado_em TEXT DEFAULT CURRENT_TIMESTAMP
                );
                CREATE TABLE ferias (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    utilizador_id INTEGER NOT NULL,
                    data_hora_inicio TEXT NOT NULL,
                    data_hora_fim TEXT NOT NULL,
                    observacao TEXT,
                    criado_em TEXT DEFAULT CURRENT_TIMESTAMP,
                    atualizado_em TEXT DEFAULT CURRENT_TIMESTAMP
                );
                """
            )
            connection.execute(
                """
                INSERT INTO utilizadores (
                    nim, posto, nome, sobrenome, data_chegada, data_partida,
                    tipo_acesso, password_salt, password_hash, master
                ) VALUES (?, 'OR-5', 'Nome', 'LEGADO', '2026-01-01 08:00',
                          '2026-12-31 20:00', 'Leitura', ?, ?, 0)
                """,
                ("legado", salt, password_hash),
            )
            connection.execute(
                """
                INSERT INTO ferias (
                    utilizador_id, data_hora_inicio, data_hora_fim, observacao
                ) VALUES (1, '2026-08-10 18:00', '2026-08-20 10:00', 'Preservar')
                """
            )
            connection.commit()
            connection.close()

            with patch.object(db, "DB_PATH", str(path)):
                db.init_db()
                columns = {row["name"] for row in db.db_rows("PRAGMA table_info(utilizadores)")}
                self.assertIn("posicao_numero", columns)
                self.assertIn("snr_substituto", columns)
                self.assertIn("snr_substituto_inicio", columns)
                self.assertIn("snr_substituto_fim", columns)
                notification_columns = {
                    row["name"]
                    for row in db.db_rows("PRAGMA table_info(ferias_notificacoes)")
                }
                self.assertIn("canal", notification_columns)
                self.assertIsNotNone(
                    db.db_one(
                        "SELECT name FROM sqlite_master "
                        "WHERE type='table' AND name='auditoria'"
                    )
                )
                period = db.db_one("SELECT * FROM ferias WHERE id=1")
                self.assertEqual("Aprovado", period["estado"])
                self.assertEqual("2026-08-10 18:00", period["data_hora_inicio"])
                self.assertEqual("2026-08-20 10:00", period["data_hora_fim"])
                self.assertIn(1, db.get_ferias_mes(2026, 8))
                history = db.db_rows(
                    "SELECT * FROM ferias_historico WHERE feria_id=1"
                )
                self.assertEqual(1, len(history))
                self.assertEqual("Migração", history[0]["acao"])


if __name__ == "__main__":
    unittest.main()
